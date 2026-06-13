"""Generate ClaimGuard prospect outreach Excel workbook."""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

from prospect_seed_data import get_prospect_seed_data  # noqa: E402

OUTPUT = ROOT / "marketing" / "ClaimGuard-Prospect-List.xlsx"

HEADERS = [
    "Business Name",
    "Sector",
    "Product Type",
    "Website",
    "Email (verify before use)",
    "Phone (verify before use)",
    "Instagram",
    "LinkedIn",
    "TikTok",
    "Reddit / Community",
    "Source",
    "Outreach Angle",
    "Status",
    "Notes",
]


def main() -> None:
    prospects = get_prospect_seed_data()
    wb = Workbook()
    ws = wb.active
    ws.title = "Prospects"

    header_fill = PatternFill("solid", fgColor="0F1729")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, prospect in enumerate(prospects, start=2):
        values = [
            prospect["business_name"],
            prospect["sector"],
            prospect["product_type"],
            prospect["website"],
            prospect["email"],
            prospect["phone"],
            prospect["instagram"],
            prospect["linkedin"],
            prospect["tiktok"],
            prospect["reddit"],
            prospect["source"],
            prospect["outreach_angle"],
            "Not contacted",
            "Verify email/phone on official website or contact page before outreach.",
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col, value=value).alignment = Alignment(wrap_text=True, vertical="top")

    widths = [28, 18, 24, 34, 28, 18, 22, 34, 22, 20, 28, 42, 14, 36]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(prospects)+1}"

    readme = wb.create_sheet("README")
    readme["A1"] = "ClaimGuard Prospect List — Important"
    readme["A1"].font = Font(name="Arial", bold=True, size=14)
    readme["A3"] = (
        "This list contains real business names and public web/social references for outreach research. "
        "Email and phone fields are intentionally blank unless publicly verified — you must confirm contacts "
        "on each brand's official website before sending campaigns. Scraping personal data from social platforms "
        "may violate platform terms and privacy laws (GDPR, CAN-SPAM, India DPDP). Use ClaimGuard outreach emails PDF "
        "with permission-based or publicly listed business contacts only."
    )
    readme["A3"].alignment = Alignment(wrap_text=True)
    readme.column_dimensions["A"].width = 110
    readme.row_dimensions[3].height = 90

    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT} ({len(prospects)} prospects)")


if __name__ == "__main__":
    main()